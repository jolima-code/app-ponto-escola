from datetime import datetime, timedelta, time
from banco_horas.models import (
    SaldoBancoHoras,
    CompensacaoBancoHoras,
)

import openpyxl

from django.contrib import messages
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db.models import Count
from django.http import HttpResponse
from django.shortcuts import render, redirect

from documentos.models import Documento
from funcionarios.models import Funcionario
from .models import RegistroPonto


def usuario_equipe(user):
    return user.is_staff


@login_required
def bater_ponto(request):
    funcionario = Funcionario.objects.get(usuario=request.user)

    ultimos_registros = RegistroPonto.objects.filter(
        funcionario=funcionario
    ).order_by("-data_hora")[:10]

    total_registros = RegistroPonto.objects.filter(
        funcionario=funcionario
    ).count()

    total_documentos = Documento.objects.filter(
        funcionario=funcionario
    ).count()

    ultimo_registro = RegistroPonto.objects.filter(
        funcionario=funcionario
    ).order_by("-data_hora").first()

    proxima_batida = "Entrada"

    if ultimo_registro:
        if ultimo_registro.tipo == "entrada":
            proxima_batida = "Saída para almoço"
        elif ultimo_registro.tipo == "saida_almoco":
            proxima_batida = "Retorno do almoço"
        elif ultimo_registro.tipo == "retorno_almoco":
            proxima_batida = "Saída"
        elif ultimo_registro.tipo == "saida":
            proxima_batida = "Entrada"

    registros_hoje = RegistroPonto.objects.filter(
        funcionario=funcionario,
        data_hora__date=datetime.now().date()
    ).order_by("data_hora")

    quantidade_batidas_hoje = registros_hoje.count()

    situacao_dia = "Presente"

    if quantidade_batidas_hoje == 0:
        situacao_dia = "Sem registro hoje"

    primeira_entrada = registros_hoje.first()
    ultima_saida = registros_hoje.last()

    horas_trabalhadas = "00:00"
    status_horas = "Normal"
    status_entrada = "Sem entrada hoje"

    horas = 0

    primeiro_registro = registros_hoje.first()

    if primeiro_registro:
        hora_entrada = primeiro_registro.data_hora.hour
        minuto_entrada = primeiro_registro.data_hora.minute

        if hora_entrada > 7 or (hora_entrada == 7 and minuto_entrada > 0):
            status_entrada = "Atrasado"
        else:
            status_entrada = "No horário"

    if quantidade_batidas_hoje >= 2:
        entrada = registros_hoje.first().data_hora
        saida = registros_hoje.last().data_hora

        diferenca = saida - entrada
        total_segundos = int(diferenca.total_seconds())

        horas = total_segundos // 3600
        minutos = (total_segundos % 3600) // 60

        horas_trabalhadas = f"{horas:02}:{minutos:02}"

    carga_diaria = float(funcionario.carga_horaria_diaria)

    if horas > carga_diaria:
        status_horas = "Hora Extra"

    if request.method == "POST":
        tipo = request.POST.get("tipo")
        observacao = request.POST.get("observacao")

        ultimo = RegistroPonto.objects.filter(
            funcionario=funcionario
        ).order_by("-data_hora").first()

        if ultimo and ultimo.tipo == tipo:
            messages.error(
                request,
                "Você já registrou esse tipo de ponto."
            )

            return redirect("bater_ponto")

        fluxo_permitido = {
            "entrada": ["saida_almoco"],
            "saida_almoco": ["retorno_almoco"],
            "retorno_almoco": ["saida"],
            "saida": ["entrada"],
        }

        if ultimo:
            proximos = fluxo_permitido.get(ultimo.tipo, [])

            if tipo not in proximos:
                messages.error(
                    request,
                    "Sequência de ponto inválida."
                )

                return redirect("bater_ponto")

        RegistroPonto.objects.create(
            funcionario=funcionario,
            tipo=tipo,
            observacao=observacao
        )

        registros_dia = RegistroPonto.objects.filter(
            funcionario=funcionario,
            data_hora__date=datetime.now().date()
        ).order_by("data_hora")

        if registros_dia.count() >= 2:

            entrada = registros_dia.first().data_hora

            saida = registros_dia.last().data_hora

            horas_trabalhadas = saida - entrada

            carga = float(
                funcionario.carga_horaria_diaria
            )

            horas_previstas = timedelta(
                hours=carga
            )

            saldo = (
                horas_trabalhadas -
                horas_previstas
            )

            SaldoBancoHoras.objects.update_or_create(

                funcionario=funcionario,

                data=datetime.now().date(),

                defaults={

                    "horas_trabalhadas": horas_trabalhadas,

                    "horas_previstas": horas_previstas,

                    "saldo": saldo,

                }

            )

        messages.success(
            request,
            "Ponto registrado com sucesso!"
        )

        return redirect("bater_ponto")

    return render(request, "ponto/bater_ponto.html", {
        "funcionario": funcionario,
        "ultimos_registros": ultimos_registros,
        "total_registros": total_registros,
        "total_documentos": total_documentos,
        "ultimo_registro": ultimo_registro,
        "proxima_batida": proxima_batida,
        "horas_trabalhadas": horas_trabalhadas,
        "status_horas": status_horas,
        "status_entrada": status_entrada,
        "quantidade_batidas_hoje": quantidade_batidas_hoje,
        "situacao_dia": situacao_dia,
        "primeira_entrada": primeira_entrada,
        "ultima_saida": ultima_saida,
        "registros_hoje": registros_hoje,
    })

@login_required
def relatorio_ponto(request):

    funcionario = Funcionario.objects.get(
        usuario=request.user
    )

    mes = request.GET.get("mes")

    registros = RegistroPonto.objects.filter(
        funcionario=funcionario
    )

    if mes:
        registros = registros.filter(
            data_hora__month=mes
        )

    registros = registros.order_by("-data_hora")

    total_registros = registros.count()

    documentos = Documento.objects.filter(
        funcionario=funcionario
    )

    if mes:
        documentos = documentos.filter(
            data_envio__month=mes
        )

    total_documentos_mes = documentos.count()

    total_tempo = timedelta()

    registros_ordenados = registros.order_by(
        "data_hora"
    )

    entrada = None

    for registro in registros_ordenados:

        if registro.tipo == "entrada":
            entrada = registro.data_hora

        elif registro.tipo == "saida" and entrada:

            total_tempo += (
                registro.data_hora - entrada
            )

            entrada = None

    total_segundos = int(
        total_tempo.total_seconds()
    )

    total_horas = total_segundos // 3600

    total_minutos = (
        total_segundos % 3600
    ) // 60

    total_horas_periodo = (
        f"{total_horas:02}:{total_minutos:02}"
    )

    return render(
        request,
        "ponto/relatorio_ponto.html",
        {
            "funcionario": funcionario,
            "registros": registros,
            "total_registros": total_registros,
            "total_horas_periodo": total_horas_periodo,
           
        }
    )

@login_required
def exportar_excel(request):
    funcionario = Funcionario.objects.get(usuario=request.user)

    registros = RegistroPonto.objects.filter(
        funcionario=funcionario
    ).order_by("-data_hora")

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Relatório de Ponto"

    sheet["A1"] = "Tipo"
    sheet["B1"] = "Data/Hora"
    sheet["C1"] = "Observação"

    linha = 2

    for registro in registros:
        sheet[f"A{linha}"] = registro.get_tipo_display()
        sheet[f"B{linha}"] = str(registro.data_hora)
        sheet[f"C{linha}"] = registro.observacao

        linha += 1

    response = HttpResponse(
        content_type="application/ms-excel"
    )

    response[
        "Content-Disposition"
    ] = 'attachment; filename="relatorio_ponto.xlsx"'

    workbook.save(response)

    return response


@login_required
@user_passes_test(usuario_equipe)
def painel_coordenacao(request):
    setor_filtro = request.GET.get("setor")

    funcionarios = Funcionario.objects.filter(
        ativo=True
    )

    if setor_filtro:
        funcionarios = funcionarios.filter(
            setor=setor_filtro
        )

    total_funcionarios = funcionarios.count()

    registros_hoje = RegistroPonto.objects.filter(
        data_hora__date=datetime.now().date()
    )

    compensacoes_pendentes = CompensacaoBancoHoras.objects.filter(
        aprovado=False
    ).order_by("-data")

    documentos_pendentes = Documento.objects.filter(
        status="pendente"
    ).count()

    documentos_aprovados = Documento.objects.filter(
        status="aprovado"
    ).count()

    documentos_rejeitados = Documento.objects.filter(
        status="rejeitado"
    ).count()

    total_registros_hoje = registros_hoje.count()

    lista_setores = Funcionario.objects.values_list(
        "setor",
        flat=True
    ).distinct()

    setores = Funcionario.objects.values(
        "setor"
    ).annotate(
        total=Count("id")
    ).order_by("setor")

    atrasados = []
    sem_registro = []
    no_horario = []
    status_colaboradores = []

    for funcionario in funcionarios:
        primeiro = RegistroPonto.objects.filter(
            funcionario=funcionario,
            data_hora__date=datetime.now().date()
        ).order_by("data_hora").first()

        status = "Sem registro"
        horario = "-"

        if primeiro:
            horario = primeiro.data_hora

            if (
                primeiro.data_hora.hour > 7 or
                (
                    primeiro.data_hora.hour == 7 and
                    primeiro.data_hora.minute > 0
                )
            ):
                status = "Atrasado"
            else:
                status = "No horário"

        status_colaboradores.append({
            "nome": funcionario.nome,
            "setor": funcionario.setor,
            "cargo": funcionario.cargo,
            "status": status,
            "horario": horario,
        })

        if not primeiro:
            sem_registro.append(funcionario)

        if primeiro:
            if (
                primeiro.data_hora.hour > 7 or
                (
                    primeiro.data_hora.hour == 7 and
                    primeiro.data_hora.minute > 0
                )
            ):
                atrasados.append(funcionario)
            else:
                no_horario.append(funcionario)

    return render(request, "ponto/painel_coordenacao.html", {
        "total_funcionarios": total_funcionarios,
        "total_registros_hoje": total_registros_hoje,
        "atrasados": atrasados,
        "sem_registro": sem_registro,
        "no_horario": no_horario,
        "setores": setores,
        "lista_setores": lista_setores,
        "setor_filtro": setor_filtro,
        "data_hoje": datetime.now().date(),
        "status_colaboradores": status_colaboradores,
        "compensacoes_pendentes": compensacoes_pendentes,
        "documentos_pendentes": documentos_pendentes,
        "documentos_aprovados": documentos_aprovados,
        "documentos_rejeitados": documentos_rejeitados,

    })


@login_required
@user_passes_test(usuario_equipe)
def exportar_coordenacao_excel(request):
    setor_filtro = request.GET.get("setor")

    funcionarios = Funcionario.objects.filter(
        ativo=True
    )

    if setor_filtro:
        funcionarios = funcionarios.filter(
            setor=setor_filtro
        )

    funcionarios = funcionarios.order_by("nome")

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Coordenação"

    sheet["A1"] = "Colaborador"
    sheet["B1"] = "Setor"
    sheet["C1"] = "Cargo"
    sheet["D1"] = "Status"

    linha = 2

    for funcionario in funcionarios:
        primeiro = RegistroPonto.objects.filter(
            funcionario=funcionario,
            data_hora__date=datetime.now().date()
        ).order_by("data_hora").first()

        status = "Sem registro"

        if primeiro:
            status = "No horário"

            if (
                primeiro.data_hora.hour > 7 or
                (
                    primeiro.data_hora.hour == 7 and
                    primeiro.data_hora.minute > 0
                )
            ):
                status = "Atrasado"

        sheet[f"A{linha}"] = funcionario.nome
        sheet[f"B{linha}"] = funcionario.setor
        sheet[f"C{linha}"] = funcionario.cargo
        sheet[f"D{linha}"] = status

        linha += 1

    response = HttpResponse(
        content_type="application/ms-excel"
    )

    response[
        "Content-Disposition"
    ] = 'attachment; filename="painel_coordenacao.xlsx"'

    workbook.save(response)

    return response
@login_required
def resumo_mensal(request):

    funcionario = Funcionario.objects.get(
        usuario=request.user
    )

    mes = request.GET.get("mes")

    registros = RegistroPonto.objects.filter(
        funcionario=funcionario
    )

    if mes:
        registros = registros.filter(
            data_hora__month=mes
        )

    total_registros = registros.count()

    total_tempo = timedelta()

    registros_ordenados = registros.order_by("data_hora")

    entrada = None

    for registro in registros_ordenados:

        if registro.tipo == "entrada":
            entrada = registro.data_hora

        elif registro.tipo == "saida" and entrada:
            total_tempo += registro.data_hora - entrada
            entrada = None

    total_segundos = int(total_tempo.total_seconds())

    total_horas = total_segundos // 3600
    total_minutos = (total_segundos % 3600) // 60

    total_horas_mes = f"{total_horas:02}:{total_minutos:02}"

    carga_mensal = float(
        funcionario.carga_horaria_diaria
    ) * 22

    horas_extras = 0

    if total_horas > carga_mensal:
        horas_extras = total_horas - carga_mensal

    dias_com_registro = registros.dates(
        "data_hora",
        "day"
    ).count()

    dias_uteis_estimados = 22

    faltas_estimadas = dias_uteis_estimados - dias_com_registro

    if faltas_estimadas < 0:
        faltas_estimadas = 0

    documentos = Documento.objects.filter(
        funcionario=funcionario
    )

    if mes:
        documentos = documentos.filter(
            data_envio__month=mes
    )

    total_documentos_mes = documentos.count()

    atrasos = registros.filter(
        tipo="entrada",
        data_hora__hour__gte=7
    ).count()

    return render(request, "ponto/resumo_mensal.html", {
        "funcionario": funcionario,
        "mes": mes,
        "total_registros": total_registros,
        "atrasos": atrasos,
        "total_documentos_mes": total_documentos_mes,
        "total_horas_mes": total_horas_mes,
        "horas_extras": horas_extras,
        "faltas_estimadas": faltas_estimadas,
        "dias_com_registro": dias_com_registro,
    })

@login_required
def meu_banco_horas(request):

    funcionario = Funcionario.objects.get(
        usuario=request.user
    )

    mes = request.GET.get("mes")

    saldos = SaldoBancoHoras.objects.filter(
        funcionario=funcionario
    )

    if mes:
        saldos = saldos.filter(
            data__month=mes
        )

    saldos = saldos.order_by("-data")

    compensacoes = CompensacaoBancoHoras.objects.filter(
        funcionario=funcionario
    ).order_by("-data")

    compensacoes_pendentes = compensacoes.filter(
        aprovado=False
    )

    compensacoes_aprovadas_lista = compensacoes.filter(
        aprovado=True
    )

    total_saldo = timedelta()

    for item in saldos:
        total_saldo += item.saldo

    compensacoes_aprovadas = compensacoes_aprovadas_lista

    for item in compensacoes_aprovadas:

        if item.tipo == "credito":
            total_saldo += item.quantidade

        elif item.tipo == "debito":
            total_saldo -= item.quantidade

        total_creditos = timedelta()
        total_debitos = timedelta()

    for item in compensacoes_aprovadas:

        if item.tipo == "credito":
            total_creditos += item.quantidade

        elif item.tipo == "debito":
            total_debitos += item.quantidade

    return render(request, "ponto/meu_banco_horas.html", {
        "funcionario": funcionario,
        "saldos": saldos,
        "total_saldo": total_saldo,
        "mes": mes,
        "compensacoes": compensacoes,
        "compensacoes_pendentes": compensacoes_pendentes,
        "compensacoes_aprovadas_lista": compensacoes_aprovadas_lista,
        "total_creditos": total_creditos,
        "total_debitos": total_debitos,
    })

@login_required
def exportar_banco_horas_excel(request):

    funcionario = Funcionario.objects.get(
        usuario=request.user
    )

    mes = request.GET.get("mes")

    saldos = SaldoBancoHoras.objects.filter(
        funcionario=funcionario
    )

    if mes:
        saldos = saldos.filter(
            data__month=mes
        )

    saldos = saldos.order_by("-data")

    workbook = openpyxl.Workbook()

    sheet = workbook.active

    sheet.title = "Banco de Horas"

    sheet["A1"] = "Data"
    sheet["B1"] = "Horas Trabalhadas"
    sheet["C1"] = "Horas Previstas"
    sheet["D1"] = "Saldo"

    linha = 2

    for item in saldos:

        sheet[f"A{linha}"] = str(item.data)

        sheet[f"B{linha}"] = str(
            item.horas_trabalhadas
        )

        sheet[f"C{linha}"] = str(
            item.horas_previstas
        )

        sheet[f"D{linha}"] = str(
            item.saldo
        )

        linha += 1

    response = HttpResponse(
        content_type="application/ms-excel"
    )

    response[
        "Content-Disposition"
    ] = (
        'attachment; '
        'filename="banco_horas.xlsx"'
    )

    workbook.save(response)

    return response

@login_required
def solicitar_compensacao(request):

    funcionario = Funcionario.objects.get(
        usuario=request.user
    )

    if request.method == "POST":

        data = request.POST.get("data")
        tipo = request.POST.get("tipo")
        quantidade = request.POST.get("quantidade")
        motivo = request.POST.get("motivo")

        horas, minutos = quantidade.split(":")

        quantidade_tempo = timedelta(
            hours=int(horas),
            minutes=int(minutos)
        )

        CompensacaoBancoHoras.objects.create(
            funcionario=funcionario,
            data=data,
            tipo=tipo,
            quantidade=quantidade_tempo,
            motivo=motivo,
            aprovado=False
        )

        messages.success(
            request,
            "Solicitação de compensação enviada para aprovação."
        )

        return redirect("meu_banco_horas")

    return render(request, "ponto/solicitar_compensacao.html", {
        "funcionario": funcionario,
    })

@login_required
@user_passes_test(usuario_equipe)
def compensacoes_coordenacao(request):

    compensacoes = CompensacaoBancoHoras.objects.all().order_by(
        "-criado_em"
    )

    pendentes = compensacoes.filter(
        aprovado=False
    )

    aprovadas = compensacoes.filter(
        aprovado=True
    )

    return render(request, "ponto/compensacoes_coordenacao.html", {
        "compensacoes": compensacoes,
        "pendentes": pendentes,
        "aprovadas": aprovadas,
    })

@login_required
@user_passes_test(usuario_equipe)
def aprovar_compensacao(request, compensacao_id):

    compensacao = CompensacaoBancoHoras.objects.get(
        id=compensacao_id
    )

    compensacao.aprovado = True

    compensacao.save()

    messages.success(
        request,
        "Compensação aprovada com sucesso."
    )

    return redirect("compensacoes_coordenacao")

@login_required
@user_passes_test(usuario_equipe)
def rejeitar_compensacao(request, compensacao_id):

    compensacao = CompensacaoBancoHoras.objects.get(
        id=compensacao_id
    )

    compensacao.delete()

    messages.success(
        request,
        "Compensação rejeitada e removida."
    )

    return redirect("compensacoes_coordenacao")

@login_required
@user_passes_test(usuario_equipe)
def documentos_coordenacao(request):

    documentos = Documento.objects.all().order_by(
        "-data_envio"
    )

    pendentes = documentos.filter(
        status="pendente"
    )

    aprovados = documentos.filter(
        status="aprovado"
    )

    rejeitados = documentos.filter(
        status="rejeitado"
    )

    return render(request, "ponto/documentos_coordenacao.html", {
        "documentos": documentos,
        "pendentes": pendentes,
        "aprovados": aprovados,
        "rejeitados": rejeitados,
    })

@login_required
@user_passes_test(usuario_equipe)
def aprovar_documento(request, documento_id):

    documento = Documento.objects.get(
        id=documento_id
    )

    documento.status = "aprovado"
    documento.save()

    messages.success(
        request,
        "Documento aprovado com sucesso."
    )

    return redirect("documentos_coordenacao")


@login_required
@user_passes_test(usuario_equipe)
def rejeitar_documento(request, documento_id):

    documento = Documento.objects.get(
        id=documento_id
    )

    documento.status = "rejeitado"
    documento.save()

    messages.success(
        request,
        "Documento rejeitado."
    )

    return redirect("documentos_coordenacao")

@login_required
@user_passes_test(usuario_equipe)
def exportar_documentos_excel(request):

    documentos = Documento.objects.all().order_by(
        "-data_envio"
    )

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Documentos"

    sheet["A1"] = "Colaborador"
    sheet["B1"] = "Título"
    sheet["C1"] = "Descrição"
    sheet["D1"] = "Status"
    sheet["E1"] = "Data de envio"

    linha = 2

    for item in documentos:

        sheet[f"A{linha}"] = item.funcionario.nome
        sheet[f"B{linha}"] = item.titulo
        sheet[f"C{linha}"] = item.descricao
        sheet[f"D{linha}"] = item.status
        sheet[f"E{linha}"] = str(item.data_envio)

        linha += 1

    response = HttpResponse(
        content_type="application/ms-excel"
    )

    response[
        "Content-Disposition"
    ] = 'attachment; filename="documentos_atestados.xlsx"'

    workbook.save(response)

    return response